import csv
import io
import os
from openpyxl import load_workbook


def load_csv_file(csv_file):
    """
    Load csv file

    :param csv_file: csv file path
    :return: list of parameters

    Examples:
        >>> cat csv_file
        environment,host,port
        ci,localhost,27017
        e2e,localhost,27018

        >>> load_csv_file(csv_file)
        [
            {'environment': 'ci', 'host': 'localhost', 'port': '27017'},
            {'environment': 'e2e', 'host': 'localhost', 'port': '27018'}
        ]

    """
    if not os.path.isabs(csv_file):
        project_working_directory = os.getcwd()
        csv_file = os.path.join(project_working_directory, *csv_file.split("/"))

    if not os.path.isfile(csv_file):
        raise FileNotFoundError

    csv_content_list = []

    with io.open(csv_file, encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            csv_content_list.append(row)

    return csv_content_list


def load_xlsx_file(xlsx_file):
    """
    Load xlsx file

    :param xlsx_file: xlsx file path
    :return: list of parameters

    """
    if not os.path.isabs(xlsx_file):
        project_working_directory = os.getcwd()
        xlsx_file = os.path.join(project_working_directory, *xlsx_file.split("/"))

    if not os.path.isfile(xlsx_file):
        raise FileNotFoundError

    wb = load_workbook(xlsx_file)
    ws = wb.active
    result_dict = []
    for row in ws.iter_rows(min_row=2):
        dict_temp = dict(zip([cell.value for cell in list(ws.iter_rows())[0]], [c.value for c in row]))
        result_dict.append(dict_temp)
    return result_dict
